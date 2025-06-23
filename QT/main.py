import sys
from PyQt6 import QtWidgets, QtCore, QtGui
from PyQt6.QtWidgets import QMessageBox, QFileDialog
from PyQt6.QtCore import QTimer, QDateTime
from ui_mainwindow import Ui_MainWindow
from influxdb_client import InfluxDBClient
from influxdb_client.client.write_api import SYNCHRONOUS
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar
from matplotlib.figure import Figure
import datetime
import pytz
import mplcursors
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill


class MonitoringApp(QtWidgets.QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)

        # Konfigurasi InfluxDB
        self.influx_url = "http://localhost:8086"
        self.influx_org = "INSTITUT TEKNOLOGI SEPULUH NOPEMBER"
        self.influx_token = "GZKhRQMga9uSiicCSmv9TliMjGNC0lakXimPMnuhvMQc3MZrmNoG339JTJc0o0A0vBnVlpA_wzp_GCrUXL6MsA=="
        self.influx_bucket = "Tank T-101"

        # Inisialisasi variabel
        self.client = None
        self.query_api = None
        self.timer = QTimer()
        self.update_interval = 10000  # 10 detik
        self.all_data = pd.DataFrame()  # Untuk menyimpan semua data
        self.temp_range = (24.0, 30.0)  # Default range suhu (min, max)
        self.humidity_range = (50.0, 70.0)  # Default range kelembaban (min, max)
        self.last_alert_time = None  # Untuk menghindari alert berulang
        self.alert_cooldown = 300  # Cooldown 5 menit (dalam detik)

        # Untuk menyimpan referensi garis chart
        self.temp_line_obj = None  # Will hold temperature Line2D object
        self.humidity_line_obj = None  # Will hold humidity Line2D object
        self.temp_cursor = None
        self.humidity_cursor = None

        # Setup UI tambahan
        self.setup_charts()
        self.setup_table()

        # Set nilai default untuk input range
        self.tempMinInput.setText(str(self.temp_range[0]))
        self.tempMaxInput.setText(str(self.temp_range[1]))
        self.humidityMinInput.setText(str(self.humidity_range[0]))
        self.humidityMaxInput.setText(str(self.humidity_range[1]))

        # Hubungkan signal
        self.startButton.clicked.connect(self.start_monitoring)
        self.stopButton.clicked.connect(self.stop_monitoring)
        self.setPointButton.clicked.connect(self.update_setpoints)
        self.exportButton.clicked.connect(self.export_to_excel)
        self.refreshButton.clicked.connect(self.refresh_table)

        # Awalnya nonaktifkan tombol stop
        self.stopButton.setEnabled(False)

    def setup_charts(self):
        """Menyiapkan grafik untuk Tab 1"""
        self.temp_figure = Figure()
        self.temp_canvas = FigureCanvas(self.temp_figure)
        self.temp_ax = self.temp_figure.add_subplot(111)
        self.temp_ax.set_title('Suhu (°C) vs Waktu')
        self.temp_ax.set_xlabel('Waktu (WIB)')
        self.temp_ax.set_ylabel('Suhu (°C)')
        self.temp_ax.grid(True)
        temp_toolbar = NavigationToolbar(self.temp_canvas, self)
        temp_layout = QtWidgets.QVBoxLayout()
        temp_layout.addWidget(temp_toolbar)
        temp_layout.addWidget(self.temp_canvas)
        self.temperatureChartView.setLayout(temp_layout)

        self.humidity_figure = Figure()
        self.humidity_canvas = FigureCanvas(self.humidity_figure)
        self.humidity_ax = self.humidity_figure.add_subplot(111)
        self.humidity_ax.set_title('Kelembaban (%) vs Waktu')
        self.humidity_ax.set_xlabel('Waktu (WIB)')
        self.humidity_ax.set_ylabel('Kelembaban (%)')
        self.humidity_ax.grid(True)
        humidity_toolbar = NavigationToolbar(self.humidity_canvas, self)
        humidity_layout = QtWidgets.QVBoxLayout()
        humidity_layout.addWidget(humidity_toolbar)
        humidity_layout.addWidget(self.humidity_canvas)
        self.humidityChartView.setLayout(humidity_layout)

    def setup_table(self):
        """Menyiapkan tabel untuk Tab 2"""
        self.tableWidget.setColumnCount(5)
        self.tableWidget.setHorizontalHeaderLabels([
            "Waktu", 
            "Lokasi", 
            "Tahap Proses", 
            "Suhu (°C)", 
            "Kelembaban (%)"
        ])
        self.tableWidget.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.ResizeMode.Stretch)
        self.tableWidget.setSortingEnabled(True)

    def update_setpoints(self):
        """Memperbarui range set point dari input pengguna"""
        try:
            temp_min = float(self.tempMinInput.text())
            temp_max = float(self.tempMaxInput.text())
            humidity_min = float(self.humidityMinInput.text())
            humidity_max = float(self.humidityMaxInput.text())
            
            # Validasi range suhu
            if temp_min >= temp_max:
                QMessageBox.warning(self, "Peringatan", "Nilai minimum suhu harus lebih kecil dari maksimum")
                return
            if temp_min < 0 or temp_max > 50:
                QMessageBox.warning(self, "Peringatan", "Suhu harus antara 0-50°C")
                return
                
            # Validasi range kelembaban
            if humidity_min >= humidity_max:
                QMessageBox.warning(self, "Peringatan", "Nilai minimum kelembaban harus lebih kecil dari maksimum")
                return
            if humidity_min < 0 or humidity_max > 100:
                QMessageBox.warning(self, "Peringatan", "Kelembaban harus antara 0-100%")
                return
                
            self.temp_range = (temp_min, temp_max)
            self.humidity_range = (humidity_min, humidity_max)
            QMessageBox.information(self, "Sukses", "Range set point berhasil diperbarui")

            # <-- TAMBAHAN: cek kondisi alert lagi berdasarkan data terbaru
            if hasattr(self, "latest_temp") and hasattr(self, "latest_humidity"):
                self.last_alert_time = None  # Reset cooldown agar alert langsung muncul
                self.check_alert_conditions(self.latest_temp, self.latest_humidity)

            
        except ValueError:
            QMessageBox.warning(self, "Peringatan", "Masukkan angka yang valid")

    def start_monitoring(self):
        """Memulai monitoring data"""
        try:
            self.client = InfluxDBClient(
                url=self.influx_url,
                token=self.influx_token,
                org=self.influx_org,
                timeout=30_000
            )
            self.query_api = self.client.query_api()

            try:
                health = self.client.health()
                if health.status == "pass":
                    self.statusLabel.setText("STATUS: Terhubung ke InfluxDB ✔")
                else:
                    self.statusLabel.setText("STATUS: Masalah Koneksi ⚠")
                    QMessageBox.warning(self, "Peringatan", f"Masalah koneksi InfluxDB: {health.message}")
                    return
            except Exception as health_error:
                self.statusLabel.setText("STATUS: Gagal Cek Kesehatan ❌")
                QMessageBox.warning(self, "Peringatan", f"Tidak bisa cek kesehatan InfluxDB: {str(health_error)}")
                return

            self.startButton.setEnabled(False)
            self.stopButton.setEnabled(True)
            self.timer.timeout.connect(self.update_data)
            self.timer.start(self.update_interval)
            self.update_data()

        except Exception as e:
            self.statusLabel.setText("STATUS: Gagal Koneksi ❌")
            QMessageBox.critical(self, "Error", f"Gagal terhubung ke InfluxDB: {str(e)}")
            if self.client:
                self.client.close()
            self.client = None
            self.query_api = None

    def stop_monitoring(self):
        """Menghentikan monitoring"""
        self.timer.stop()
        if self.client:
            self.client.close()
            self.client = None
            self.query_api = None

        self.statusLabel.setText("STATUS: Terputus ⛔")
        self.startButton.setEnabled(True)
        self.stopButton.setEnabled(False)
        QMessageBox.information(self, "Info", "Monitoring dihentikan")

    def check_alert_conditions(self, temp_value, humidity_value):
        """Memeriksa kondisi alert berdasarkan range set point"""
        current_time = datetime.datetime.now()
        
        # Skip jika masih dalam cooldown
        if self.last_alert_time and (current_time - self.last_alert_time).total_seconds() < self.alert_cooldown:
            return
            
        messages = []
        
        # Check temperature conditions with more detailed messages
        if temp_value < self.temp_range[0]:
            diff = self.temp_range[0] - temp_value
            messages.append(
                f"⚠ PERINGATAN SUHU RENDAH ⚠\n"
                f"Suhu saat ini: {temp_value:.1f}°C\n"
                f"Range normal: {self.temp_range[0]}-{self.temp_range[1]}°C\n"
                f"(Terlalu rendah {diff:.1f}°C)\n"
                f"Dampak: Viskositas meningkat (produk mengental, sulit dipompa)"
            )
        elif temp_value > self.temp_range[1]:
            diff = temp_value - self.temp_range[1]
            messages.append(
                f"⚠ PERINGATAN SUHU TINGGI ⚠\n"
                f"Suhu saat ini: {temp_value:.1f}°C\n"
                f"Range normal: {self.temp_range[0]}-{self.temp_range[1]}°C\n"
                f"(Terlalu tinggi {diff:.1f}°C)\n"
                f"Dampak: Overpressure → Risiko ledakan."
            )
            
        # Check humidity conditions with more detailed messages
        if humidity_value < self.humidity_range[0]:
            diff = self.humidity_range[0] - humidity_value
            messages.append(
                f"⚠ PERINGATAN KELEMBABAN RENDAH ⚠\n"
                f"Kelembaban saat ini: {humidity_value:.1f}%\n"
                f"Range normal: {self.humidity_range[0]}-{self.humidity_range[1]}%\n"
                f"(Terlalu rendah {diff:.1f}%)\n"
                f"Dampak: Peningkatan risiko static electricity (bahaya percikan api)"
            )
        elif humidity_value > self.humidity_range[1]:
            diff = humidity_value - self.humidity_range[1]
            messages.append(
                f"⚠ PERINGATAN KELEMBABAN TINGGI ⚠\n"
                f"Kelembaban saat ini: {humidity_value:.1f}%\n"
                f"Range normal: {self.humidity_range[0]}-{self.humidity_range[1]}%\n"
                f"(Terlalu tinggi {diff:.1f}%)\n"
                f"Dampak: PPertumbuhan mikroba (bakteri pengurai hidrokarbon)"
            )
            
        # Tampilkan alert jika ada pesan
        if messages:
            self.last_alert_time = current_time
            alert_message = "\n\n".join(messages)
            
            # Create a more prominent warning dialog
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Icon.Warning)
            msg.setWindowTitle("PERINGATAN KONDISI LINGKUNGAN")
            msg.setText(alert_message)
            msg.setStandardButtons(QMessageBox.StandardButton.Ok)
            
            # Make the dialog larger
            msg.setStyleSheet("QLabel{min-width: 500px; min-height: 150px;}")
            msg.exec()

    def update_data(self):
        """Memperbarui data dari InfluxDB"""
        if not self.query_api:
            self.statusLabel.setText("STATUS: Query API tidak tersedia ❌")
            QMessageBox.warning(self, "Peringatan", "Query API belum diinisialisasi")
            return

        try:
            query = f'''
            from(bucket: "{self.influx_bucket}")
              |> range(start: -24h) 
              |> filter(fn: (r) => r["_measurement"] == "environment_monitoring")
              |> filter(fn: (r) => r["_field"] == "humidity_percent" or r["_field"] == "temperature_celsius")
              |> filter(fn: (r) => r["location"] == "Crude Oil Storage Tank T-101")
              |> filter(fn: (r) => r["process_stage"] == "Storage")
              |> filter(fn: (r) => r["sensor_id"] == "SHT20-001")
              |> filter(fn: (r) => exists r._value)  
              |> yield(name: "raw")  
            '''

            try:
                result = self.query_api.query(query)
            except Exception as query_error:
                self.statusLabel.setText("STATUS: Error Query ⚠")
                QMessageBox.warning(self, "Error Query", f"Gagal menjalankan query: {str(query_error)}")
                return

            temp_data = []
            humidity_data = []
            temp_times = []
            humidity_times = []
            records_list = []  # Untuk menyimpan data ke tabel
            latest_temp = None
            latest_humidity = None

            for table in result:
                for record in table.records:
                    if record.get_field() == "temperature_celsius":
                        temp_data.append(record.get_value())
                        temp_times.append(record.get_time())
                        latest_temp = record.get_value()
                    elif record.get_field() == "humidity_percent":
                        humidity_data.append(record.get_value())
                        humidity_times.append(record.get_time())
                        latest_humidity = record.get_value()

                    if not self.locationLabel.text().startswith("LOCATION:"):
                        self.locationLabel.setText(f"LOKASI: {record.values.get('location', 'N/A')}")
                        self.processStageLabel.setText(f"PROSES: {record.values.get('process_stage', 'N/A')}")
                        self.sensorIdLabel.setText(f"SENSOR ID: {record.values.get('sensor_id', 'N/A')}")

                    # Simpan data untuk tabel
                    records_list.append({
                        'time': record.get_time(),
                        'location': record.values.get('location', 'N/A'),
                        'process_stage': record.values.get('process_stage', 'N/A'),
                        'field': record.get_field(),
                        'value': record.get_value()
                    })

            if temp_data and temp_times:
                self.update_chart(self.temp_ax, self.temp_canvas, temp_times, temp_data, 'Suhu (°C)')
            if humidity_data and humidity_times:
                self.update_chart(self.humidity_ax, self.humidity_canvas, humidity_times, humidity_data, 'Kelembaban (%)')

            # Perbarui data tabel
            if records_list:
                self.update_data_table(records_list)

            # Check alert conditions
            if latest_temp is not None and latest_humidity is not None:
                self.check_alert_conditions(latest_temp, latest_humidity)

            self.latest_temp = latest_temp
            self.latest_humidity = latest_humidity

            now = QDateTime.currentDateTime()
            self.updateLabel.setText(f"Terakhir Diperbarui: {now.toString('dd MMMM yyyy - hh:mm:ss')}")

        except Exception as e:
            self.statusLabel.setText("STATUS: Error Pembaruan ⚠")
            QMessageBox.warning(self, "Error", f"Error memperbarui data: {str(e)}")

    def update_chart(self, ax, canvas, times, values, title):
        """Memperbarui grafik dengan data baru"""
        try:
            ax.clear()
            local_tz = pytz.timezone('Asia/Jakarta')
            local_times = [t.astimezone(local_tz) for t in times]
            line, = ax.plot(local_times, values, 'b-')
            
            # Tambahkan garis range set point
            if title == 'Suhu (°C)':
                range_min, range_max = self.temp_range
                if self.temp_cursor:
                    self.temp_cursor.remove()
                self.temp_line_obj = line  # Store line object
            else:
                range_min, range_max = self.humidity_range
                if self.humidity_cursor:
                    self.humidity_cursor.remove()
                self.humidity_line_obj = line  # Store line object
                
            # Garis untuk range minimum dan maksimum
            ax.axhline(y=range_min, color='r', linestyle='--', label='Range Min')
            ax.axhline(y=range_max, color='r', linestyle='--', label='Range Max')
            
            # Area fill antara range
            ax.fill_between(local_times, range_min, range_max, color='green', alpha=0.1)
            
            ax.legend()
            ax.set_title(title)
            ax.set_xlabel('Waktu (WIB)')
            ax.set_ylabel(title.split(' ')[0])
            ax.grid(True)
            ax.xaxis.set_major_formatter(plt.matplotlib.dates.DateFormatter('%H:%M:%S', tz=local_tz))
            plt.setp(ax.get_xticklabels(), rotation=45)
            formatted_times = [t.strftime('%Y-%m-%d %H:%M:%S') for t in local_times]
            
            if title == 'Suhu (°C)':
                self.temp_cursor = mplcursors.cursor(line, hover=True)
                def on_add_temp(sel):
                    idx = sel.target.index
                    sel.annotation.set_text(
                        f"{title.split(' ')[0]}: {values[idx]:.2f}\nWaktu: {formatted_times[idx]}"
                    )
                self.temp_cursor.connect("add", on_add_temp)
            else:
                self.humidity_cursor = mplcursors.cursor(line, hover=True)
                def on_add_humidity(sel):
                    idx = sel.target.index
                    sel.annotation.set_text(
                        f"{title.split(' ')[0]}: {values[idx]:.2f}\nWaktu: {formatted_times[idx]}"
                    )
                self.humidity_cursor.connect("add", on_add_humidity)
            canvas.draw()

        except Exception as e:
            QMessageBox.warning(self, "Error Grafik", f"Error memperbarui grafik: {str(e)}")

    def update_data_table(self, new_records):
        """Memperbarui tabel dengan data baru"""
        try:
            # Konversi records ke DataFrame
            new_df = pd.DataFrame(new_records)
            
            # Gabungkan dengan data yang sudah ada
            if not self.all_data.empty:
                # Gabungkan dan hapus duplikat
                self.all_data = pd.concat([self.all_data, new_df]).drop_duplicates(
                    subset=['time', 'field'], 
                    keep='last'
                )
            else:
                self.all_data = new_df
            
            # Pivot data untuk tampilan tabel
            df_pivot = self.all_data.pivot_table(
                index=['time', 'location', 'process_stage'], 
                columns='field', 
                values='value'
            ).reset_index()
            
            # Konversi waktu ke timezone lokal
            local_tz = pytz.timezone('Asia/Jakarta')
            df_pivot['time'] = pd.to_datetime(df_pivot['time']).dt.tz_convert(local_tz)
            
            # Format waktu untuk tampilan
            df_pivot['time_str'] = df_pivot['time'].dt.strftime('%Y-%m-%d %H:%M:%S')
            
            # Simpan data lengkap untuk ekspor
            self.export_data = df_pivot.copy()
            
            # Perbarui tabel
            self.refresh_table()

        except Exception as e:
            QMessageBox.warning(self, "Error Tabel", f"Error memperbarui tabel: {str(e)}")

    def refresh_table(self):
        """Memperbarui tampilan tabel dengan data terbaru"""
        try:
            if hasattr(self, 'export_data') and not self.export_data.empty:
                # Nonaktifkan sorting sementara untuk performa
                self.tableWidget.setSortingEnabled(False)
                
                # Set jumlah baris sesuai data yang ada
                self.tableWidget.setRowCount(len(self.export_data))
                
                # Isi data ke dalam tabel baris per baris
                for row_idx, row in self.export_data.iterrows():
                    # Buat item untuk setiap kolom
                    waktu_item = QtWidgets.QTableWidgetItem(row['time_str'])
                    lokasi_item = QtWidgets.QTableWidgetItem(row['location'])
                    proses_item = QtWidgets.QTableWidgetItem(row['process_stage'])
                    
                    # Ambil nilai suhu dan kelembaban
                    suhu = row.get('temperature_celsius', None)
                    kelembaban = row.get('humidity_percent', None)
                    
                    # Format teks untuk nilai suhu dan kelembaban
                    teks_suhu = f"{suhu:.2f}" if suhu is not None else "N/A"
                    teks_kelembaban = f"{kelembaban:.2f}" if kelembaban is not None else "N/A"
                    
                    # Buat item tabel untuk suhu dan kelembaban
                    suhu_item = QtWidgets.QTableWidgetItem(teks_suhu)
                    kelembaban_item = QtWidgets.QTableWidgetItem(teks_kelembaban)
                    
                    # Masukkan item ke dalam tabel
                    self.tableWidget.setItem(row_idx, 0, waktu_item)
                    self.tableWidget.setItem(row_idx, 1, lokasi_item)
                    self.tableWidget.setItem(row_idx, 2, proses_item)
                    self.tableWidget.setItem(row_idx, 3, suhu_item)
                    self.tableWidget.setItem(row_idx, 4, kelembaban_item)
                    
                    # Periksa apakah nilai di luar setpoint
                    if suhu is not None and kelembaban is not None:
                        suhu_diluar = suhu < self.temp_range[0] or suhu > self.temp_range[1]
                        kelembaban_diluar = kelembaban < self.humidity_range[0] or kelembaban > self.humidity_range[1]
                        
                        # Jika ada nilai yang di luar setpoint, warnai baris dengan merah
                        if suhu_diluar or kelembaban_diluar:
                            for kolom in range(self.tableWidget.columnCount()):
                                item = self.tableWidget.item(row_idx, kolom)
                                if item:
                                    # Gunakan warna merah muda untuk highlight
                                    item.setBackground(QtGui.QColor(255, 200, 200))
                
                # Aktifkan kembali fitur sorting
                self.tableWidget.setSortingEnabled(True)
                
        except Exception as e:
            # Tampilkan pesan error jika terjadi masalah
            QMessageBox.warning(self, "Error", f"Gagal memperbarui tabel: {str(e)}")

    def export_to_excel(self):
        """Export data to Excel with color formatting"""
        if not hasattr(self, 'export_data') or self.export_data.empty:
            QMessageBox.warning(self, "Peringatan", "Tidak ada data untuk diekspor")
            return
            
        try:
            file_name, _ = QFileDialog.getSaveFileName(
                self, "Simpan File Excel", "", 
                "File Excel (*.xlsx);;Semua File (*)")
            
            if not file_name:
                return
                
            if not file_name.endswith('.xlsx'):
                file_name += '.xlsx'
            
            # Prepare data
            export_df = self.export_data.copy()
            if 'time' in export_df.columns:
                export_df['time'] = export_df['time'].dt.tz_localize(None)
            
            export_df = export_df[['time', 'location', 'process_stage', 
                                 'temperature_celsius', 'humidity_percent']]
            export_df.columns = ['Waktu', 'Lokasi', 'Tahap Proses', 
                              'Suhu (°C)', 'Kelembaban (%)']
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            
            # Define highlight color
            highlight_fill = PatternFill(
                start_color='FFFFC8C8', 
                end_color='FFFFC8C8',
                fill_type='solid'
            )
            
            # Add headers
            ws.append(list(export_df.columns))
            
            # Add data rows
            for _, row in export_df.iterrows():
                ws.append(list(row))
                
                # Check for out-of-range values
                temp = row['Suhu (°C)']
                humidity = row['Kelembaban (%)']
                
                if pd.notna(temp) and pd.notna(humidity):
                    temp_out = temp < self.temp_range[0] or temp > self.temp_range[1]
                    hum_out = humidity < self.humidity_range[0] or humidity > self.humidity_range[1]
                    
                    if temp_out or hum_out:
                        for cell in ws[ws.max_row]:
                            cell.fill = highlight_fill
            
            wb.save(file_name)
            QMessageBox.information(self, "Sukses", 
                                  "Data berhasil diekspor ke Excel dengan format warna")
            
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Gagal mengekspor data: {str(e)}")

def main():
    app = QtWidgets.QApplication(sys.argv)
    window = MonitoringApp()
    window.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()